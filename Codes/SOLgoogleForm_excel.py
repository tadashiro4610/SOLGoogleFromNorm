from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
import openpyxl as op
from openpyxl.styles.alignment import Alignment
import re

class sfg:
    def __init__(self,dir_open,dir_save,col_dir=None):
        

        self.wb_open=op.load_workbook(dir_open)
        self.ws_open=self.wb_open.worksheets[0]
        
        self.wb_save=op.load_workbook(dir_save)
        self.ws_save=self.wb_save.worksheets[0]
        
        self.wb_col_open=op.load_workbook(col_dir)
        self.ws_col_open=self.wb_col_open.worksheets[0]
    # ##デバッグ用
    # def __init__(self):
    #     self.wb_open=op.load_workbook("c:/Users/tadas/Downloads/八王子祭エントリー.xlsx")
    #     self.ws_open=self.wb_open.worksheets[0]
    #     self.wb_save=op.load_workbook("test1.xlsx")
    #     self.ws_save=self.wb_save.worksheets[0]
    # ##デバッグ用

        #初期化
        no_borders = op.styles.borders.Border(
                top = None,
                bottom = None,
                left = None,
                right = None,
            )
        for row in self.ws_save:
            for cell in row:
                cell.value = None
                cell.fill = PatternFill(fill_type = None)
                cell.border = no_borders

        # self.bandName=[] #バンド名
        self.nameDic={} #何バンド加入してるかの辞書
        self.member=[] #バンドメンバー記述用
        # self.songCount=[] #曲数
        self.canNOTAct=[] #出演できない日
        maxlen=0

        ###辞書の更新用関数(ない場合は追加、ある場合はvalueを+1)
        def update_dictionary(dictionary, key):
            if key in dictionary:
                dictionary[key] += 1
            else:
                dictionary[key] = 1
        
        ###(1,1)->A1に変換する関数
        def cellNum(row,col):
            return self.ws_save.cell(row=row,column=col).coordinate
        
        ###シートの数値と色塗りをクリア
        for row in self.ws_save:
            for cell in row:
                cell.value = None
                cell.fill=PatternFill(fill_type = None)

        ###バンド名
        #読み取って書込み
        i=2
        self.ws_save[cellNum(1,1)]="バンド名"
        for cols in self.ws_open.iter_cols(min_row=2, min_col=2, max_row=self.ws_open.max_row, max_col=2):
            for cell in cols:
                #書込み
                num=cellNum(i,1)
                self.ws_save[num]=cell.value
                i+=1

        ###曲数
        #読み取って書込み
        i=2
        self.ws_save[cellNum(1,2)]="曲数"
        for cols in self.ws_open.iter_cols(min_row=2, min_col=4, max_row=self.ws_open.max_row, max_col=4):
            for cell in cols:
                #書込み
                # print(cell.value)
                # val=int(cell.value)
                num=cellNum(i,2)
                # print(re.findall(r"\d+",val)[0])
                # self.ws_save[num]=re.findall(r"\d+",val)[0]
                self.ws_save[num]=cell.value
                i+=1
        
        ###メンバー読み取り(名前のみ)
        #読み取って、辞書に追加して、1バンドごとに書込み
        i=2
        self.ws_save[cellNum(1,3)]="メンバー"
        for cols in self.ws_open.iter_cols(min_row=2, min_col=3, max_row=self.ws_open.max_row, max_col=3):
            for cell in cols:
                self.member=[]
                # print(cell.value)
                split_txt=cell.value.splitlines() #改行で区切る(名前/学年/パートになるはず)
                for k in range(len(split_txt)):
                    split=re.split("[/,]",split_txt[k]) #/で区切って最初を取得
                    update_dictionary(self.nameDic,split[0]) #辞書にキーがなかったら追加、有ったら+1(血反吐度)
                    self.member.append(split[0])

                maxlen=max(maxlen,len(self.member)) #最大のバンドメンバー数
                #書込み
                for j in range (3,len(self.member)+3):
                    num=cellNum(i,j)
                    self.ws_save[num]=self.member[j-3]
                i+=1
        
        ###出演可能日
        #読み取って書込み
        i=2
        self.ws_save[cellNum(1,maxlen+3)]="出演可能日"
        for cols in self.ws_open.iter_cols(min_row=2, min_col=5, max_row=self.ws_open.max_row, max_col=5):
            for cell in cols:
                #書込み
                # print(cell.value)
                num=cellNum(i,maxlen+3)
                self.ws_save[num]=cell.value
                i+=1

        ###出演不可時間
        i=2
        self.ws_save[cellNum(1,maxlen+4)]="出演不可時間"
        for cols in self.ws_open.iter_cols(min_row=2, min_col=6, max_row=self.ws_open.max_row, max_col=6):
            for cell in cols:
                #書込み
                # print(cell.value)
                num=cellNum(i,maxlen+4)
                self.ws_save[num]=cell.value
                i+=1

        ###コメント
        i=2
        self.ws_save[cellNum(1,maxlen+5)]="コメント"
        for cols in self.ws_open.iter_cols(min_row=2, min_col=7, max_row=self.ws_open.max_row, max_col=7):
            for cell in cols:
                #書込み
                # print(cell.value)
                num=cellNum(i,maxlen+5)
                self.ws_save[num]=cell.value
                i+=1

        ###メンバーの色分け(血反吐化)
        #セルの値を読み取って、辞書参照、色塗り
        i=2
        max_key=max(self.nameDic.values())
        print(max_key)
        hex_color=[]
        if col_dir==None:
            for p in range(255, 0, -((255-0)//max_key)):
                # RGB配列で指定
                color = (255, p, p)
                # RBG配列を16進数に変換(openpyxlの仕様上16進数のみ対応)
                hex_color.append(str.format('{:02x}{:02X}{:02X}', color[0], color[1], color[2]) )
            # print(hex_color)
        else:
            for cols in self.ws_col_open.iter_cols(min_row=1, min_col=1, max_row=self.ws_open.max_row, max_col=1):
                for p in cols:
                    if p.fill.fgColor.value=="00000000":
                        hex_color.append("FFFFFFFF")
                    else:
                        hex_color.append(p.fill.fgColor.value)
            print(hex_color)
            
        
        for cols in self.ws_open.iter_cols(min_row=2, min_col=3, max_row=self.ws_open.max_row, max_col=3):
            for cell in cols:
                for j in range (3,maxlen+3):
                    num=cellNum(i,j)
                    # print(num)
                    self.wb_save.save(dir_save)
                    #辞書のvalueによって血反吐度を上げていく     
                    for q in range(1,max_key+1):
                        if self.ws_save[num].value==None:
                            pass
                        elif self.nameDic[self.ws_save[num].value]==q:
                            self.ws_save[num].fill=PatternFill(patternType='solid', fgColor=hex_color[q-1])
                            print(hex_color[q-1])
                i+=1
        
        ###出演者と出演数の表示
        self.k=[]
        self.v=[]
        #タプルからkeyとvalueをリスト化
        for k, v in self.nameDic.items():
            self.k.append(k)
            self.v.append(v)
        #列名
        self.ws_save[cellNum(self.ws_open.max_row+2,1)]="名前"
        self.ws_save[cellNum(self.ws_open.max_row+2,2)]="出演数"
        #書込み
        for i in range(0,len(self.k)):
            self.ws_save[cellNum(self.ws_open.max_row+3+i,1)]=self.k[i]
            self.ws_save[cellNum(self.ws_open.max_row+3+i,2)]=self.v[i]

        ###フォームのテンプレの挿入
        self.ws_save.insert_rows(1,2)
        self.ws_save[(cellNum(1,1))]="フォームテンプレ"
        self.ws_save[(cellNum(1,2))]="https://docs.google.com/forms/d/1FshCvDuC6LhjPHlEO8-sbQxwuVP9TvTQB59qhg5yZdU/edit"
        self.wb_save.save(dir_save)

# a=sfg() #デバッグ用